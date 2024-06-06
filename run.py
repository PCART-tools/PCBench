import os
import ast
import time
import random
import subprocess
from multiprocessing import Pool
from multiprocessing import Manager
from openpyxl import load_workbook
random.seed(42)
from utils import *

#这里的paraStr都是不带最外层的括号的
def projPara2Obj(paraStr):
    paraObjLst=[] #保存参数对象
    lst=[]
    if paraStr:
        lst=get_parameter(paraStr,space=0) #项目参数不去空格
    for i in range(len(lst)):
        para=lst[i]
        parameter=Parameter()
        parameter.position=i
        parameter.fullItem=para
        if '=' in para:
            pos=para.find('=')
            if '(' not in para[0:pos] and "'" not in para[0:pos] and '"' not in para[0:pos]: #等号前面也不能出现引号，比如f('x=2',y=1)
                parameter.name=para[0:pos].replace(' ','')
                parameter.value=para[pos+1:]
            else:
               parameter.value=para
        else:
            parameter.value=para
        
        paraObjLst.append(parameter)
    
    return paraObjLst
    

#根据用户传递的参数值，给库中定义的每个参数都添加上参数值
def addValue(libPara,projPara):
    for it in libPara:
        flag=0
        for itt in projPara: #优先按照参数名确定库与项目中参数的对应关系
            if it.name==itt.name:
                it.value=itt.value
                flag=1
                break
        
        if flag==0:
            for itt in projPara: #若根据名字找不到对应关系，再根据位置找对应关系
                if it.position==itt.position:
                    it.value=itt.value
                    break
    

#n: 不带默认值的位置参数个数
#N: 位置参数的总个数
#m: 带默认值的关键字参数的个数
def calculate(n,N,m):
    sum=0
    for i in range(n,N+1): #[n,N]中的参数都是带默认值的
        if m:
            sum+=(i+1)*(1+m+m-1) #1个都不挑+m中各挑一个+依次增加m中的个数
        else:
            sum+=(i+1)

    return sum


def generateCombinations(libPara,projPara):
    libPosLst,libKeyLst=para2Obj(libPara)
    projParaLst=projPara2Obj(projPara)
    indexPos=0 #记录最后一个不带默认值的位置参数下标,极端的情况是位置参数都带默认值，则indexPos就是0
    for it in libPosLst: #找出不带默认值的位置参数
        if not it.value:
            indexPos+=1
    
    temp=[]
    indexKey=0
    for it in libKeyLst: #找出不带默认值的关键字参数
        if not it.value:
            temp.append(it.name)
            indexKey+=1

    #给库参数填上参数值
    addValue(libPosLst,projParaLst)
    addValue(libKeyLst,projParaLst)
    keyNoDefaultValue=[] #保存不带默认值的关键字参数（不带默认值的参数在组合的时候是必须要填的）
    keyWithDefaultValue=[] #保存带默认值的关键字参数
    
    for it in libKeyLst:
        flag=1
        for itt in temp: #temp中保存的是不带默认值的关键字参数
            if it.name==itt:
                keyNoDefaultValue.append(f"{it.name}={it.value}")
                flag=0
                break
        if flag:
            keyWithDefaultValue.append(f"{it.name}={it.value}")
    
    # 计算将会产生多少组合
    # num=calculate(indexPos,len(libPosLst),len(keyWithDefaultValue))
    # print(f"posNoValue={indexPos}, posNum={len(libPosLst)}, keyWithValue={len(keyWithDefaultValue)}, combinations={num}")
    
    ansLst=[]
    if len(libPosLst)>0:
        while indexPos<=len(libPosLst): 
            #step1: 改变参数个数,按照位置顺序，依次增加位置参数的个数
            lst=libPosLst[0:indexPos] #不带默认值的位置参数一定要出现在组合中
            
            #step2:修改位置参数的使用方式，即从后往前逐个给位置参数加上参数名
            lst.reverse() 
            for pos in range(len(lst)+1):
                paraWithName=[]
                for it in lst[0:pos]:
                    paraWithName.append(f"{it.name}={it.value}")
                paraWithName.reverse() #加上参数名后再逆序回来
                paraWithName+=keyNoDefaultValue
                
                paraNoName=[]
                for it in lst[pos:]:
                    paraNoName.append(it.value)
                paraNoName.reverse()

                #先把位置参数固定到组合中
                ansLst.append(', '.join(paraNoName+paraWithName))

                #step3将带默认值的关键参数依次加到组合中，再对其打乱顺序
                #一次只挑一个带默认值的关键字参数,确保每个参数有单独使用的机会
                for para in keyWithDefaultValue:
                    tempLst=paraWithName+[para]
                    random.shuffle(tempLst) #随机改变顺序
                    ansLst.append(', '.join(paraNoName+tempLst))
                
                #依次增加带默认值的参数 
                for i in range(2,len(keyWithDefaultValue)+1):
                    tempLst=paraWithName+keyWithDefaultValue[0:i]
                    random.shuffle(tempLst) #随机改变顺序
                    ansLst.append(', '.join(paraNoName+tempLst))
                    

            indexPos+=1
    
    else:
        #先把不带默认值的关键字参数提前固定到组合中
        if len(keyNoDefaultValue)>0:
            tempLst=keyNoDefaultValue
            random.shuffle(tempLst) #随机打乱顺序
            ansLst.append(', '.join(tempLst))
        else: #如果全是带默认值的关键字参数则可以一个都不挑
            ansLst.append('')

        #一次只挑一个带默认值的
        for para in keyWithDefaultValue:
            tempLst=keyNoDefaultValue+[para]
            random.shuffle(tempLst)
            ansLst.append(', '.join(tempLst))
        
        #依次增加带默认值的关键字参数 
        for i in range(2,len(keyWithDefaultValue)+1):
            tempLst=keyNoDefaultValue+keyWithDefaultValue[0:i]
            random.shuffle(tempLst)
            ansLst.append(', '.join(tempLst))


    newLst=list(set(ansLst)) #去重
    newLst.sort(key=ansLst.index)
    # with open("组合.txt",'w') as fw:
    #     for it in newLst:
    #         fw.write(f"({it})\n")
    return newLst 






def readExcel(filePath,libName):
    wb=load_workbook(filePath)
    sheet=wb[libName] #通过名字获取sheet
    ans=[]
    for row in sheet.iter_rows(min_row=2): #每个row是一个tuple,包含一行中的所有元素,从第2行开始读,第一行是标题
        if row[0].value==None:
            break
        apiName=row[0].value.replace(' ','') #第一列的元素是API名字
        currentVerison=row[1].value
        targetVersion=row[2].value
        currentDefinition=row[3].value
        targetDefinition=row[4].value
        codeSnippet=row[5].value #API用例
        changeType=row[6].value
        ans.append((apiName,currentVerison,targetVersion,currentDefinition,targetDefinition,codeSnippet,changeType))
    wb.close()
    return ans



def getProjPara(apiName,root):
    for node in ast.walk(root):
        if isinstance(node,ast.Call):
            callName=ast.unparse(node.func)
            temp=removeParameter(callName)
            # if apiName in callName or callName in apiName:
            if apiName==temp.split('.')[-1]:
                argLst=[]
                for arg in node.args:
                    argLst.append(ast.unparse(arg))
                for keyword in node.keywords:
                    argLst.append(ast.unparse(keyword))
                parameters=','.join(argLst)
                return parameters
    return None



def modifyParameter(codeSnippet,apiName,parameters):
    root=ast.parse(codeSnippet,filename='<unknown>',mode='exec')
    paraObj=projPara2Obj(parameters)
    for node in ast.walk(root):
        if isinstance(node,ast.Call):
            callName=ast.unparse(node.func)
            #去除callName中的所有参数
            temp=removeParameter(callName)
            # if apiName in callName or callName in apiName:
            if apiName==temp.split('.')[-1]:
                posLst=[]
                keyLst=[]
                for para in paraObj:
                    if para.name:
                        paraNode=ast.keyword(arg=para.name,value=ast.Name(id=para.value))
                        keyLst.append(paraNode)
                    else:
                        paraNode=ast.Name(id=para.value)
                        posLst.append(paraNode)
                node.args=posLst
                node.keywords=keyLst
    
    newCode=ast.unparse(root)
    return newCode





def buildBench(args):
    Tuple,lock,libName=args
    apiName=Tuple[0].replace(' ','') #去重API名中的空格
    currentVersion=Tuple[1]
    targetVersion=Tuple[2]
    currentDefinition=Tuple[3]
    codeSnippet=Tuple[5]

    #判断代码片段是否解析失败 
    try:
        root=ast.parse(codeSnippet,filename='<unknown>',mode='exec')
    except Exception as e:
        with lock:
            print(f"{apiName}: ast.parse failed,{e}")
            print(f"{codeSnippet}\n")
            with open('error.txt','a') as fe:
                fe.write(f"{apiName}@{currentVersion}, parse to ast failed")
        return
    
    #step1: 让测试用例在起始版本运行，来判断测试用例是否生成成功
    #在这一步就开始插桩，看它是否被覆盖到
    codeSnippet=ast.unparse(root) #预处理代码片段，把多行改写到一行 
    codeLst=codeSnippet.splitlines()
    api=apiName.split('.')[-1]
    for i in range(len(codeLst)):
        if api in codeLst[i] and 'def ' not in codeLst[i]:
            spaceNum=countSpace(codeLst[i])
            s1="print('debug_in')\n"
            s2="print('debug_out')\n"
            while spaceNum>0:
                s1=' '+s1
                s2=' '+s2
                spaceNum-=1
            s=s1+codeLst[i]+'\n'+s2
            codeLst[i]=s
            break
    tempFile=f"example/{apiName}@{currentVersion}-{targetVersion}.py"
    with open(tempFile,'w') as fw:
        for it in codeLst:
            fw.write(f"{it}\n")
    pythonPath=f"/dataset/zhang/anaconda3/envs/{libName}{currentVersion}/bin/python"
    command=f'{pythonPath} "{tempFile}"'
    result=subprocess.run(command,shell=True,executable='/bin/bash',stdout=subprocess.PIPE,stderr=subprocess.PIPE,text=True)
    
    if result.returncode!=0:
        with lock:
            print(f"{apiName} run failed at curren version{currentVersion}")
            os.remove(tempFile)
            print(result.stderr)
            return
    elif not ('debug_in' in result.stdout and 'debug_out' in result.stdout):
        with lock:
            print(f"{apiName}@{currentVersion} uncoverd")
            os.remove(tempFile)
            return
    
    os.remove(tempFile)
    projPara=getProjPara(api,root) #获取API在项目中的真实参数
    if projPara==None:
        with lock:
            print(f"{apiName} name error")
        return 
    combs=generateCombinations(currentDefinition,projPara)
    
    if len(combs)==0:
        combs.append('')
    try:
        os.mkdir(f"Temp/{libName}")
    except FileExistsError:
        pass

    if not os.path.exists(f"Temp/{libName}/{apiName}@{currentVersion}-{targetVersion}"):
        os.mkdir(f"Temp/{libName}/{apiName}@{currentVersion}-{targetVersion}")
    
    start=time.perf_counter()
    failCnt=0
    count=0  #记录在当前版本能够成功运行的测试用例
    for i in range(len(combs)):
        codeText=modifyParameter(codeSnippet,api,combs[i])#采用ast来替换参数，正则匹配容易出问题

        tempFile=f"{apiName}@{currentVersion}-{targetVersion}#{i+1}.py"
        with open(tempFile,'w') as fw:
            fw.write(f"{codeText}\n")

        pythonPath=f"/dataset/zhang/anaconda3/envs/{libName}{currentVersion}/bin/python"
        command=f'{pythonPath} "{tempFile}"'
        
        #让测试用例在起始版本运行，来判断测试用例是否生成成功
        result=subprocess.run(command,shell=True,executable='/bin/bash',stdout=subprocess.PIPE,stderr=subprocess.PIPE,text=True)
        if result.returncode!=0:
            failCnt+=1
            with lock:
                print("check error.txt")
                with open('error.txt','a') as fe:
                    fe.write(f"\n{apiName}#{i+1} runs falied: {result.stderr}\n")
                    fe.write(f"{codeText}\n")
                    continue
        
        
        #先让测试用例在目标版本运行，不能运行，则肯定不兼容
        count+=1
        os.remove(tempFile)
        dirName=f"Temp/{libName}/{apiName}@{currentVersion}-{targetVersion}/{apiName}#{count}"
        if not os.path.exists(dirName):
            os.mkdir(dirName)
        with open(f"{dirName}/{api}.py",'w') as fw:
            fw.write(f"{codeText}\n")
        
    end=time.perf_counter()
    with lock:
        with open("time.txt",'a') as fw:
            fw.write(f"{apiName}: run time={int(end-start)}s, success={len(combs)-failCnt}, fail={failCnt}\n")



def main():
    #step1:读取Excel文件名
    libLst=['faker', 'fastapi', 'click', 'django', 'xgboost', 'dask', 'aiohttp', 'keras', 'gensim', 'spacy', 'requests', 
            'tornado', 'httpx', 'transformers', 'pandas', 'tensorflow', 'matplotlib', 'plotly', 'sympy', 'pydantic', 'flask', 'torch', 
            'numpy', 'polars', 'networkx', 'redis', 'jax', 'scipy', 'PIL', 'sklearn', 'loguru', 'lightgbm', 'rich']

    for libName in libLst[0:2]:
        lst=readExcel('ChangedAPI.xlsx',libName)
        manager=Manager()
        lock=manager.Lock() #创建一个共享的锁
        tasks=[(Tuple,lock,libName) for Tuple in lst] #一个进程处理一个API 
        pool=Pool(processes=15)
        pool.map(buildBench,tasks)
        pool.close()
        pool.join() #等待进程池中所有的任务执行完，否则主进程可能继续往下执行提前结束，而导致部分任务没有执行完


if __name__=='__main__':
    main()